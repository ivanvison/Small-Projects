# Purpose of this code is to read one XLSX file that has a lot of data in particular columns and rows, pick them up and add them to specific column in a masterfile
# This was made using Python and ChatGPT. The next tweak was to have the script read all the files in a folder and add them to the masterfile. 


import os
from openpyxl import load_workbook

file_path = r"I:\Path-to-Folder\With-Files\FORMAT 1\3.xlsx"
master_file_path = r"I:\Path-to-Folder\With-Files\masterfile.xlsx"

def find_company_row_in_master(master_sheet, company_name):
    for row_number, row in enumerate(master_sheet.iter_rows(values_only=True), start=1):
        if row[0] == company_name:
            return row_number
    return None

def check_in_master_file(company_name):
    try:
        master_wb = load_workbook(filename=master_file_path, data_only=True)
        master_sheet = master_wb.active

        # Check if the company name is in column A of the master file
        row_number = find_company_row_in_master(master_sheet, company_name)
        if row_number is not None:
            total_amount = master_sheet.cell(row=row_number, column=2).value
            return True, total_amount, row_number

        return False, None, None
    except Exception as e:
        print(f"Error occurred while checking in master file: {e}")
        return False, None, None

def update_master_file(company_name, total_amount, row_number):
    try:
        master_wb = load_workbook(filename=master_file_path)
        master_sheet = master_wb.active

        # Update the total amount in column XxXx for the corresponding company name
        master_sheet.cell(row=row_number, column=59, value=total_amount) 

        master_wb.save(master_file_path)
    except Exception as e:
        print(f"Error occurred while updating master file: {e}")

def process_excel_file(file_path):
    wb = load_workbook(filename=file_path, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]

    company_names = []
    total_amounts = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        company_name = row[1]
        total_amount = row[20]

        if isinstance(company_name, str) and company_name.startswith('Total ') and total_amount is not None:
            company_names.append(company_name.replace('Total ', ''))
            total_amounts.append(total_amount)

    with open(r"I:\Path-to-Folder\With-Files\output.txt", 'w') as output_file:
        output_file.write("Company Name - Total Amount - In Masterfile\n")

        for company, amount in zip(company_names, total_amounts):
            is_in_masterfile, _, row_number = check_in_master_file(company)
            output_file.write(f"{company} - {amount} - {'' if is_in_masterfile else 'No'}\n")
    
            if is_in_masterfile:
                update_master_file(company, amount, row_number)

def main():
    process_excel_file(file_path)

if __name__ == "__main__":
    main()
